import openpyxl
import pymysql
import os

def import_student_data():
    file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data/student_info.xlsx")
    # print(file_path)
    workbook_data = openpyxl.load_workbook(file_path)

    work_sheet = workbook_data["G2306"]

    max_row = work_sheet.max_row
    print("===================================================", work_sheet, max_row - 1) # 打印工作表
    start_row = 2
    cell_data_arr = []
    while start_row <= max_row:
        cell_data_班级 = work_sheet.cell(row=start_row, column=2)
        cell_data_姓名 = work_sheet.cell(row=start_row, column=3)
        cell_data_性别 = work_sheet.cell(row=start_row, column=4)
        cell_data_民族 = work_sheet.cell(row=start_row, column=5)
        cell_data_身份证 = work_sheet.cell(row=start_row, column=6)
        cell_data_地址 = work_sheet.cell(row=start_row, column=8)

        cell_data = {
            "clazz": cell_data_班级.value,
            "name": cell_data_姓名.value,
            "gender": 1 if cell_data_性别.value == "男"  else 0,
            "nation": cell_data_民族.value,
            "idcard": cell_data_身份证.value,
            "grade": "高一"
        }
        cell_data_arr.append(cell_data)
        start_row = start_row + 1
    
    print(cell_data_arr)
    return cell_data_arr

def insert_into_database(arr):
    # 创建连接  
    conn = pymysql.connect(host="mysql.sqlpub.com", port=3306, user='sylvie233', passwd='Wikdyuora4HqNZQg', db='ygtz_jiaowu_test', charset='utf8mb4')  
    
    # 创建游标(查询数据返回为字典格式)  
    cursor = conn.cursor(pymysql.cursors.DictCursor)  

    for it in arr:
    # 1. 执行SQL,返回受影响的行数  
        effect_row1 = cursor.execute(f"insert into ytgz_student(name, gender, grade, clazz, nation, idcard) values(\"{it['name']}\", \"{it['gender']}\", \"{it['grade']}\", \"{it['clazz']}\", \"{it['nation']}\", \"{it['idcard']}\")")  

    # 增/删/改均需要进行commit提交,进行保存  
    conn.commit()  
    
    # 关闭游标  
    cursor.close()  
    # 关闭连接  
    conn.close()  


if __name__ == "__main__":
    result = import_student_data()
    insert_into_database(result)
